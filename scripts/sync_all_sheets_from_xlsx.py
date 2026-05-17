#!/usr/bin/env python3
"""
Sync books-data.js from 民族志信息表-*.xlsx using ALL sheets (union of rows).

- Rows dedupe on (norm_author, year, full English title from 《cn》(en)).
- When the same key appears in multiple sheets, keep the longer summary; tie-break prefers 总表.
- Match existing JS books with fuzzy English subtitles (table uses full titles; JS may use short).
- Preserves id, lat, lon, sites on matched books; updates title, author, year, publisher, summary,
  sourceField, locationEn, and location/countryOrRegion when book has no `sites`.
- Rows with no matching book: append with new id; coordinates from rough country centroids when needed.
"""
from __future__ import annotations

import json
import re
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
BOOKS_PATH = ROOT / "books-data.js"
PREFIX = "window.ETHNOGRAPHY_BOOKS = "

# Prefer 总表 when merging duplicate keys; then first-seen sheet order.
SHEET_PRIORITY = {
    "总表": 0,
    "东亚": 1,
    "东南亚": 2,
    "亚马逊": 3,
    "赞米亚": 4,
    "美拉尼西亚": 5,
    "欧洲": 6,
    "中东": 7,
    "非洲": 8,
    "阿拉伯地区": 9,
    "北非": 10,
    "伊朗-中亚": 11,
}


def norm_title_cell(t: str) -> str:
    t = str(t or "").replace("（", "(").replace("）", ")")
    return re.sub(r"\s+", " ", t.strip())


def parse_cn_en_title(raw_title):
    text = norm_title_cell(raw_title)
    m = re.match(r"^《\s*(.*?)\s*》\s*\(\s*(.*?)\s*\)\s*$", text)
    if not m:
        return None, None
    return m.group(1).strip(), m.group(2).strip()


def canonical_title(cn: str, en: str) -> str:
    return f"《{cn}》({en})"


def norm_author(raw) -> str:
    a = str(raw or "").strip()
    a = a.replace("（", "(").replace("）", ")")
    a = re.sub(r"\([^)]*\)\s*$", "", a).strip()
    return re.sub(r"\s+", " ", a)


# Present in spreadsheets but deliberately omitted from books-data.js.
_SPREADSHEET_EXCLUDES: frozenset[tuple[str, int, str]] = frozenset(
    {
        ("魏明毅", 2016, norm_title_cell("Quiet Workers: Days and Nights at the Dock")),
    },
)

# Editorial wording fixed in-repo; xlsx still has older phrasing — re-apply after each row merge.
_SUMMARY_OVERRIDE_BY_BOOK_ID: dict[str, str] = {
    "ethnography-45": (
        "中医如何在上海与旧金山之间的跨国流通中，被重新组装为全球替代医学与中国现代性的双重象征。\n\n"
        "上海中医院与旧金山中医实践者的多点民族志调查。\n\n"
        "中医是\"异质世界化\"的产物，既非传统的遗存，也非西方的他者；其跨国流通揭示了现代性本身的多元路径。"
    ),
    "ethnography-93": (
        "贝鲁特南郊什叶派穆斯林女性如何将公开虔诚实践与对现代性的渴望融合，形成独特的\"附魅主体\"——既非世俗自由主义，也非保守传统主义。"
        "  达希亚区女性宗教活动、纪念仪式与日常生活的长期民族志调查。"
        "  现代性与宗教虔诚并非天然对立；什叶派女性通过公开表演的虔诚，在伊斯兰框架内争取主体性，有力挑战了西方女性主义对能动性的狭隘预设。"
    ),
}


# Chinese display title tweaks not yet reflected in the spreadsheet canonical cell.
_TITLE_OVERRIDE_BY_BOOK_ID: dict[str, str] = {
    "ethnography-06": "《牧灵诊所》(The Pastoral Clinic)",
    "ethnography-78": "《全球\"猎身\"》(Global \"Body Shopping\")",
    "ethnography-132": "《抵抗的灵魂和资本主义规训》(Spirits of Resistance and Capitalist Discipline: Factory Women in Malaysia)",
}

def spreadsheet_row_excluded(row: dict) -> bool:
    return (
        norm_author(row.get("author", "")),
        int(row["year"]),
        norm_title_cell(row.get("en", "")),
    ) in _SPREADSHEET_EXCLUDES


def parse_country_location(raw) -> tuple[str, str]:
    s = (raw or "")
    s = s.replace("\r\n", "\n").strip()
    if not s:
        return "", ""
    lines = [ln.strip() for ln in s.split("\n") if ln.strip()]
    if len(lines) >= 2:
        return lines[0], lines[1]
    one = lines[0]
    if re.search(r"[\u4e00-\u9fff]", one) and re.search(r"[A-Za-z]", one):
        for m in re.finditer(r"\s+(?=[A-Za-z])", one):
            left, right = one[: m.start()].strip(), one[m.start() :].strip()
            if re.search(r"[\u4e00-\u9fff]", left) and re.search(r"[A-Za-z]", right):
                if not re.search(r"[\u4e00-\u9fff]", right):
                    return left, right
    return one, ""


def normalize_location_display(value: str) -> str:
    s = str(value or "").strip()
    # Use · only for country-location joins; keep lexical hyphens in English names intact.
    s = re.sub(r"(?<![A-Za-z])\s*-\s*(?![A-Za-z])", "·", s)
    s = re.sub(r"\s*·\s*", "·", s)
    s = re.sub(r"\s*/\s*", " / ", s)
    return re.sub(r"\s{2,}", " ", s).strip()


def en_to_location_en(en_line: str) -> str:
    if not en_line:
        return ""
    return normalize_location_display(en_line)


def en_to_country_location(en_line: str) -> tuple[str, str]:
    if not en_line.strip():
        return "", ""
    if ", " in en_line and " - " not in en_line:
        return en_line.strip(), ""
    parts = re.split(r"\s+-\s+", en_line, maxsplit=1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return en_line.strip(), ""


def norm_en_for_match(s: str) -> str:
    s = norm_title_cell(s).lower()
    s = re.sub(r"^the\s+", "", s)
    return re.sub(r"[^\w\s]", "", s)


def en_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    na, nb = norm_en_for_match(a), norm_en_for_match(b)
    if na == nb:
        return 1.0
    if na in nb or nb in na:
        return 0.92
    return SequenceMatcher(None, na, nb).ratio()


def titles_match_en(en_js: str, en_xlsx: str) -> bool:
    if en_js.strip() == en_xlsx.strip():
        return True
    if en_similarity(en_js, en_xlsx) >= 0.78:
        return True
    # Short JS title is prefix of longer xlsx (e.g. "Friction" vs full subtitle).
    ej, ex = en_js.strip().lower(), en_xlsx.strip().lower()
    if ex.startswith(ej + ":") or ex.startswith(ej + " ") or ex.startswith(ej + "("):
        return True
    return False


def en_same_book_edition_variant(a_en: str, b_en: str) -> bool:
    """True when spreadsheet rows cite the same work with short/long subtitle variants."""
    if not (a_en and b_en):
        return False
    aa, bb = norm_title_cell(a_en), norm_title_cell(b_en)
    if aa.lower() == bb.lower():
        return True
    sa, sb = norm_en_for_match(aa), norm_en_for_match(bb)
    if sa == sb:
        return True
    shorter, longer = (aa.lower(), bb.lower()) if len(aa) <= len(bb) else (bb.lower(), aa.lower())
    if longer.startswith(shorter.rstrip()) and len(longer) > len(shorter):
        suffix = longer[len(shorter) :].lstrip()
        if suffix.startswith(":") or suffix.startswith(" ("):
            return True
    return SequenceMatcher(None, sa, sb).ratio() >= 0.97


def merge_spreadsheet_variant_rows(rows: list[dict]) -> dict:
    primary = sorted(
        rows,
        key=lambda r: (
            -len(r["en"]),
            -(len(str(r.get("summary") or "").strip())),
            row_sort_key(r.get("sheet", "")),
        ),
    )[0]
    base = dict(primary)
    for other in rows:
        if other is primary:
            continue
        summ_o = str(other.get("summary") or "").strip()
        if len(summ_o) > len(str(base.get("summary") or "").strip()):
            base["summary"] = other["summary"]
        pub_o = str(other.get("publisher") or "").strip()
        if pub_o and len(pub_o) > len(str(base.get("publisher") or "").strip()):
            base["publisher"] = other["publisher"]
        loc_o = other.get("country_loc")
        if (
            isinstance(loc_o, str)
            and len(loc_o.strip()) > len(str(base.get("country_loc") or "").strip())
        ):
            base["country_loc"] = loc_o
    return base


def collapse_edition_duplicate_rows(rows: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, int], list[dict]] = {}
    for r in rows:
        grouped.setdefault((norm_author(r["author"]), int(r["year"])), []).append(r)
    folded: list[dict] = []
    for group in grouped.values():
        if len(group) == 1:
            folded.extend(group)
            continue
        indices = list(range(len(group)))
        remaining = set(indices)
        while remaining:
            i = remaining.pop()
            bag_idxs = [i]
            progressed = True
            while progressed:
                progressed = False
                removable = []
                for j in list(remaining):
                    if any(
                        en_same_book_edition_variant(group[j]["en"], group[k]["en"])
                        for k in bag_idxs
                    ):
                        bag_idxs.append(j)
                        removable.append(j)
                        progressed = True
                for j in removable:
                    remaining.discard(j)
            chunk = [group[j] for j in bag_idxs]
            folded.append(chunk[0] if len(chunk) == 1 else merge_spreadsheet_variant_rows(chunk))

    folded.sort(key=lambda d: (d["author"].lower(), d["year"], d["en"].lower()))
    return folded


def row_sort_key(sheet_name: str) -> int:
    return SHEET_PRIORITY.get(sheet_name, 99)


def load_merged_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    buckets: dict[tuple[str, int, str], list[tuple[int, str, dict]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            title = ws.cell(r, 1).value
            author = ws.cell(r, 2).value
            year = ws.cell(r, 3).value
            publisher = ws.cell(r, 4).value
            country_loc = ws.cell(r, 5).value
            summary = ws.cell(r, 6).value
            if not (title and author and year):
                continue
            cn, en = parse_cn_en_title(title)
            if not cn or not en:
                raise SystemExit(f"Bad title in sheet {sheet_name!r} row {r}: {title!r}")
            au = norm_author(author)
            yr = int(year)
            key = (au, yr, en.strip())
            summ = str(summary or "").strip()
            pub = str(publisher or "").strip()
            rowdict = {
                "title_raw": title,
                "cn": cn,
                "en": en.strip(),
                "author": str(author).strip(),
                "year": yr,
                "publisher": pub,
                "country_loc": country_loc,
                "summary": summ,
                "sheet": sheet_name,
            }
            buckets.setdefault(key, []).append((row_sort_key(sheet_name), sheet_name, rowdict))

    out: list[dict] = []
    for key, items in buckets.items():
        items.sort(key=lambda x: x[0])
        # Merge: base on lowest priority index; longest summary wins field-by-field.
        base = dict(items[0][2])
        for _, _, other in items[1:]:
            if len(other.get("summary") or "") > len(base.get("summary") or ""):
                base["summary"] = other["summary"]
            if (other.get("publisher") or "") and len(other["publisher"]) > len(base.get("publisher") or ""):
                base["publisher"] = other["publisher"]
        out.append(base)

    out.sort(key=lambda d: (d["author"].lower(), d["year"], d["en"].lower()))
    return collapse_edition_duplicate_rows(out)


def load_books():
    content = BOOKS_PATH.read_text(encoding="utf-8")
    if not content.startswith(PREFIX):
        raise ValueError("books-data.js format changed: missing expected prefix")
    payload = content[len(PREFIX) :].strip()
    if payload.endswith(";"):
        payload = payload[:-1]
    return json.loads(payload)


def save_books(books):
    output = PREFIX + json.dumps(books, ensure_ascii=False, indent=2) + ";\n"
    BOOKS_PATH.write_text(output, encoding="utf-8")


def book_en_title(book: dict) -> str:
    cn, en = parse_cn_en_title(book.get("title", ""))
    if en:
        return en
    return norm_title_cell(book.get("title", ""))


def next_ethnography_id(books):
    best = 0
    for b in books:
        m = re.match(r"^ethnography-(\d+)$", str(b.get("id", "")))
        if m:
            best = max(best, int(m.group(1)))
    return f"ethnography-{best + 1:02d}"


# Rough primary pin when a book is new to JS (country/region from English location line).
ROUGH_LAT_LON: dict[str, tuple[float, float]] = {
    "china": (35.0, 103.0),
    "japan": (36.2, 138.25),
    "south korea": (37.55, 126.98),
    "indonesia": (-2.5, 118.0),
    "malaysia": (4.2, 101.97),
    "vietnam": (21.03, 105.85),
    "thailand": (15.87, 100.99),
    "india": (20.59, 78.96),
    "bangladesh": (23.68, 90.35),
    "pakistan": (30.38, 69.35),
    "iran": (32.43, 53.69),
    "iraq": (33.31, 44.37),
    "israel": (31.77, 35.21),
    "palestine": (31.95, 35.23),
    "lebanon": (33.89, 35.5),
    "syria": (34.8, 38.99),
    "yemen": (15.55, 48.52),
    "kuwait": (29.31, 47.48),
    "saudi arabia": (23.89, 45.08),
    "egypt": (26.82, 30.8),
    "morocco": (31.79, -7.09),
    "algeria": (28.03, 1.66),
    "tunisia": 33.886917,  # typo - fix below
    "nigeria": 9.0765,  # fix
}

# Fix ROUGH_LAT_LON typos - use proper tuples
ROUGH_LAT_LON = {
    "china": (35.0, 103.0),
    "japan": (36.2, 138.25),
    "south korea": (37.55, 126.98),
    "indonesia": (-2.5, 118.0),
    "malaysia": (4.2, 101.97),
    "vietnam": (21.03, 105.85),
    "thailand": (15.87, 100.99),
    "india": (20.59, 78.96),
    "bangladesh": (23.68, 90.35),
    "pakistan": (30.38, 69.35),
    "iran": (32.43, 53.69),
    "iraq": (33.31, 44.37),
    "israel": (31.77, 35.21),
    "palestine": (31.95, 35.23),
    "lebanon": (33.89, 35.5),
    "cyprus": (35.13, 33.43),
    "kuwait": (29.31, 47.48),
    "oman": (21.47, 55.92),
    "turkey": (38.96, 35.24),
    "yemen": (15.55, 48.52),
    "afghanistan": (33.94, 67.71),
    "uzbekistan": (41.38, 64.59),
    "kyrgyzstan": (41.2, 74.77),
    "tajikistan": (38.86, 71.28),
    "central asia": (42.0, 68.0),
    "iraq / lebanon": (33.9, 36.2),
    "cameroon": (7.37, 12.35),
    "côte d'ivoire": (7.54, -5.55),
    "cote d'ivoire": (7.54, -5.55),
    "sierra leone": (8.46, -11.79),
    "nigeria": (9.08, 8.68),
    "egypt": (26.82, 30.8),
    "algeria": (28.03, 1.66),
    "morocco": (31.79, -7.09),
    "brazil": (-14.24, -51.93),
    "mexico": (23.63, -102.55),
    "usa": (39.83, -98.58),
    "united states": (39.83, -98.58),
    "chile": (-35.68, -71.54),
    "canada": (56.13, -106.35),
    "uk": (55.378, -3.436),
    "united kingdom": (55.378, -3.436),
    "france": (46.23, 2.21),
    "italy": (41.87, 12.57),
    "romania": (45.94, 24.97),
    "russia": (61.52, 105.32),
    "soviet union": (61.52, 105.32),
    "ukraine": (48.38, 31.17),
    "netherlands": (52.13, 5.29),
    "botswana": (-22.33, 24.68),
    "papua new guinea": (-6.31, 143.96),
    "taiwan": (23.65, 120.975),
    "madagascar": (-19.0, 46.7),
    "finland": (66.504, 25.729),
    "australia": (-25.27, 133.78),
    "colombia": (4.57, -74.3),
    "ecuador": (-1.83, -78.18),
    "venezuela": (6.42, -66.59),
    "peru": (-9.19, -75.02),
    "haiti": (18.97, -72.29),
    "global multi-site": (20.0, 0.0),
    "west africa": (8.0, -3.0),
    "atlantic africa": (8.0, 1.0),
    "europe": (50.11, 9.8),
    "bosnia-herzegovina": (43.92, 17.68),
    "democratic republic of the congo": (-4.04, 21.76),
    "drc": (-4.04, 21.76),
    "serbia": (44.02, 20.91),
    "south africa": (-30.56, 22.94),
    "spain": (40.46, -3.75),
    "togo": (8.62, 0.82),
    "burkina faso": (12.24, -1.56),
    "north america": (45.5, -100.0),
    "israel / palestine": (31.77, 35.21),
}


def guess_lat_lon(en_line: str) -> tuple[float, float]:
    cor, _loc = en_to_country_location(en_line)
    if not cor and en_line:
        parts = re.split(r"[/,]", en_line)
        cor = parts[0].strip() if parts else ""
    key = cor.strip().lower()
    # Strip parenthetical qualifiers for lookup
    key = re.sub(r"\s*\([^)]*\)\s*", "", key).strip()
    if key in ROUGH_LAT_LON:
        return ROUGH_LAT_LON[key]
    if "/" in key:
        first_region = key.split("/", maxsplit=1)[0].strip()
        if first_region in ROUGH_LAT_LON:
            return ROUGH_LAT_LON[first_region]
    # First token
    first = key.split()[0] if key else ""
    if first in ROUGH_LAT_LON:
        return ROUGH_LAT_LON[first]
    return (20.0, 0.0)


EXACT_COORDS_BY_EN_TITLE: dict[str, tuple[float, float]] = {
    "Everything Was Forever, Until It Was No More": (59.9343, 30.3351),
    "Dreams that Matter": (30.0444, 31.2357),
    "The Moral Neoliberal": (45.4642, 9.19),
    "Friction": (-0.7893, 113.9213),
    "The Intimate Economies of Bangkok": (13.7563, 100.5018),
    "Everyday Conversions": (29.3759, 47.9774),
    "Global Body Shopping": (17.385, 78.4867),
    "Nostalgia for the Future": (6.1725, 1.2314),
    "Sonic Socialism": (21.0278, 105.8342),
    "The War Machines": (8.484, -13.2299),
    "When Bodies Remember": (-26.2041, 28.0473),
    "Enforcing Order": (48.901, 2.45),
    "Kinshasa": (-4.4419, 15.2663),
    "The Biopolitics of Beauty": (-22.9068, -43.1729),
    "Marginal Gains": (6.335, 5.6037),
    "The Land of Open Graves": (31.3322, -110.9378),
    "After the Revolution": (44.7866, 20.4489),
    "Creative Reckonings": (30.0444, 31.2357),
    # Pernambuco sugar zone / Timbaúba area (Northeast Brazil field site).
    "Death Without Weeping": (-7.5053, -35.3181),
    "The Anxieties of Mobility": (1.1301, 104.0529),
    "Shaving the Beasts": (42.55, -8.42),
    "The Reckoning of Pluralism": (41.0082, 28.9784),
    "The Vanishing Hectare": (45.91, 23.27),
    "An Enchanted Modern": (33.85, 35.52),
    "Peripheral Visions": (15.3694, 44.191),
    "Border Work": (40.5, 71.0),
    "In the Time of Oil": (22.967, 57.299),
    "The Underneath of Things": (7.956, -11.74),
    "The New Woman in Uzbekistan": (40.3864, 71.7864),
    "Evicted from Eternity": (41.895, 12.492),
    # Peninsular Malaysia: shared pin for works whose field sites are close (Western/Central Malay states).
    "Islamic Modern": (4.399, 101.53),
    "Bazaar Politics": (34.83, 69.08),
    "Ungovernable Life": (33.3152, 44.3661),
    "Passionate Uprisings": (35.6892, 51.389),
    "Algeria in France": (48.8566, 2.3522),
    "The Perils of Belonging": (3.848, 11.5021),
    "The Reindeer People": (65.0, 145.0),
    "Memories of the Slave Trade": (9.0, -12.0),
    "Warring Souls": (35.6892, 51.389),
    "Political Spiritualities": (6.5244, 3.3792),
    "Waste Siege": (31.9466, 35.3027),
    "Yearnings in the Meantime": (43.8563, 18.4131),
    "Knot of the Soul": (33.5731, -7.5898),
    "The Performance of Human Rights in Morocco": (33.5731, -7.5898),
    "The Will to Improve": (-2.5489, 120.324),
    "The Make-Believe Space": (35.1856, 33.3823),
    "On the Edge of the Global": (-21.18, -175.2),
    # New / edge-case titles merged from spreadsheets (short/long subtitle variants resolve here).
    "Lost People": (-19.87, 46.75),
    "Lost People: Magic and the Legacy of Slavery in Madagascar": (-19.87, 46.75),
    "Global Cinderellas": (25.033, 121.565),
    "Hunters, Pastoralists and Ranchers": (66.504, 25.729),
    "Spirits of Resistance and Capitalist Discipline": (4.399, 101.53),
    "Spirits of Resistance and Capitalist Discipline: Factory Women in Malaysia": (4.399, 101.53),
}


EXACT_SITES_BY_EN_TITLE: dict[str, list[dict[str, float]]] = {
    "Global Body Shopping": [
        {"lat": 17.385, "lon": 78.4867},
        {"lat": -33.8688, "lon": 151.2093},
    ],
    "The War Machines": [
        {"lat": 8.484, "lon": -13.2299},
        {"lat": 6.3156, "lon": -10.8074},
    ],
    "Ungovernable Life": [
        {"lat": 33.3152, "lon": 44.3661},
        {"lat": 33.8938, "lon": 35.5018},
    ],
    "The Perils of Belonging": [
        {"lat": 3.848, "lon": 11.5021},
        {"lat": 6.8276, "lon": -5.2893},
    ],
}


def apply_row_to_book(book: dict, row: dict) -> None:
    zh_loc, en_loc = parse_country_location(row["country_loc"])
    loc_en = en_to_location_en(en_loc)
    title_canon = canonical_title(row["cn"], row["en"])
    book["title"] = title_canon
    bid = str(book.get("id") or "")
    if bid in _TITLE_OVERRIDE_BY_BOOK_ID:
        book["title"] = _TITLE_OVERRIDE_BY_BOOK_ID[bid]
    book["author"] = str(row["author"]).strip()
    book["year"] = row["year"]
    book["publisher"] = row["publisher"] or book.get("publisher", "")
    book["summary"] = row["summary"]
    if bid in _SUMMARY_OVERRIDE_BY_BOOK_ID:
        book["summary"] = _SUMMARY_OVERRIDE_BY_BOOK_ID[bid]
    book["sourceField"] = zh_loc or book.get("sourceField", "")
    book["locationEn"] = loc_en or book.get("locationEn", "")
    cor, loc = en_to_country_location(en_loc)
    if cor:
        book["countryOrRegion"] = cor
    book["location"] = loc if loc else cor
    if not book.get("sites"):
        if not cor:
            book["location"] = book.get("location", "")
    exact = EXACT_COORDS_BY_EN_TITLE.get(row["en"])
    if exact:
        book["lat"], book["lon"] = exact
        if row["en"] in EXACT_SITES_BY_EN_TITLE:
            book["sites"] = EXACT_SITES_BY_EN_TITLE[row["en"]]
        else:
            book.pop("sites", None)
    elif not book.get("sites") and (
        book.get("lat") is None or book.get("lon") is None or (book.get("lat"), book.get("lon")) == (20.0, 0.0)
    ):
        book["lat"], book["lon"] = guess_lat_lon(en_loc)
    if len(book.get("sites") or []) <= 1:
        book.pop("sites", None)


def exact_sites_for_title(en_title: str) -> list[dict[str, float]] | None:
    sites = EXACT_SITES_BY_EN_TITLE.get(en_title)
    if sites is None:
        return None
    return [dict(site) for site in sites]


def create_book_from_row(row: dict, new_id: str) -> dict:
    zh_loc, en_loc = parse_country_location(row["country_loc"])
    loc_en = en_to_location_en(en_loc)
    cor, loc = en_to_country_location(en_loc)
    lat, lon = guess_lat_lon(en_loc)
    if row["en"] in EXACT_COORDS_BY_EN_TITLE:
        lat, lon = EXACT_COORDS_BY_EN_TITLE[row["en"]]
    book = {
        "id": new_id,
        "title": canonical_title(row["cn"], row["en"]),
        "year": row["year"],
        "author": str(row["author"]).strip(),
        "publisher": row["publisher"],
        "summary": row["summary"],
        "location": loc if loc else cor,
        "countryOrRegion": cor,
        "lat": lat,
        "lon": lon,
        "sourceField": zh_loc,
        "locationEn": loc_en,
    }
    sites = exact_sites_for_title(row["en"])
    if sites:
        book["sites"] = sites
    return book


def main():
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument(
        "xlsx",
        nargs="?",
        default=str(ROOT / "民族志信息表-0513.xlsx"),
        help="Path to workbook",
    )
    args = ap.parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        raise SystemExit(f"Missing xlsx: {xlsx_path}")

    rows = load_merged_rows(xlsx_path)
    rows_before = len(rows)
    rows = [r for r in rows if not spreadsheet_row_excluded(r)]
    excluded_rows = rows_before - len(rows)
    books = load_books()

    used_book_ids: set[str] = set()
    matched_rows: set[tuple[str, int, str]] = set()
    new_books: list[dict] = []
    appended: list[dict] = []

    for row in rows:
        au, yr, en_x = norm_author(row["author"]), row["year"], row["en"]
        candidates = [
            b
            for b in books
            if b.get("id") not in used_book_ids
            and norm_author(b.get("author")) == au
            and int(b.get("year", -1)) == yr
        ]
        best = None
        best_score = 0.0
        en_js_prev = None
        for b in candidates:
            en_js = book_en_title(b)
            if titles_match_en(en_js, en_x):
                sc = en_similarity(en_js, en_x)
                if sc > best_score:
                    best_score = sc
                    best = b
                    en_js_prev = en_js

        if best:
            apply_row_to_book(best, row)
            used_book_ids.add(str(best["id"]))
            matched_rows.add((au, yr, en_x))
            new_books.append(best)
        else:
            book = create_book_from_row(row, next_ethnography_id(books + appended))
            appended.append(book)
            new_books.append(book)

    # Append-only path should be empty if spreadsheet ⊂ old JS; user sheet has 112 vs 118 — need orphans?
    remaining_rows = []
    # Re-scan: rows that failed - actually we require every row matched. Unmatched xlsx impossible if raise.
    # Old books not in spreadsheet:
    removed = [b for b in books if str(b["id"]) not in used_book_ids]

    def id_num(b):
        m = re.match(r"^ethnography-(\d+)$", str(b.get("id", "")))
        return int(m.group(1)) if m else 9999

    new_books.sort(key=id_num)

    save_books(new_books)
    print(f"xlsx_merged_rows: {rows_before}")
    print(f"xlsx_after_exclusions: {len(rows)} (excluded: {excluded_rows})")
    print(f"books_after: {len(new_books)}")
    print(f"appended_books: {len(appended)}")
    for b in appended:
        print(f"  + {b.get('id')} {b.get('title')} | {b.get('author')} {b.get('year')}")
    print(f"removed_books: {len(removed)}")
    for b in sorted(removed, key=id_num):
        print(f"  - {b.get('id')} {b.get('title')} | {b.get('author')} {b.get('year')}")


if __name__ == "__main__":
    main()
